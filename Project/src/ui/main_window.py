import sys
from PySide6.QtWidgets import (QMainWindow, QTableWidget, QVBoxLayout, QWidget,
							   QComboBox, QHBoxLayout, QLabel, QHeaderView,
							   QTableWidgetItem, QFrame, QPushButton, QButtonGroup,
							   QDialog, QMessageBox, QFileDialog)
from PySide6.QtCore import Qt, QFile, QIODevice
from PySide6.QtGui import QColor, QFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


class ExcelLoader:
	@staticmethod
	def load_excel_data(file_path):
		"""Загружает данные из Excel файла и возвращает структурированные данные"""
		workbook = openpyxl.load_workbook(file_path)
		data = {}

		# Обрабатываем первую форму (бухгалтерский баланс)
		sheet1 = workbook['форма1']
		balance_data = ExcelLoader.process_balance_sheet(sheet1)
		data['balance'] = balance_data

		# Обрабатываем вторую форму (отчет о прибылях и убытках)
		sheet2 = workbook['форма2']
		profit_loss_data = ExcelLoader.process_profit_loss_sheet(sheet2)
		data['profit_loss'] = profit_loss_data

		return data

	@staticmethod
	def process_balance_sheet(sheet):
		"""Обрабатывает лист с бухгалтерским балансом"""
		data = []
		years = ['2013', '2014', '2015']

		# Парсим данные из таблицы
		for row in sheet.iter_rows(min_row=20, values_only=True):
			if not any(row):  # Пропускаем пустые строки
				continue

			if row[0] and isinstance(row[0], str) and row[0].startswith('I.') or row[0].startswith('II.'):
				# Это заголовок раздела
				data.append({'section': row[0].strip()})
			elif row[0] and isinstance(row[0], str) and row[2] and isinstance(row[2], int):
				# Это строка с данными
				item = {
					'parameter': row[0].strip(),
					'code': str(row[2]),
					'2013': float(row[4]) if row[4] and row[4] != '-' else 0,
					'2014': float(row[5]) if row[5] and row[5] != '-' else 0,
					'2015': float(row[3]) if row[3] and row[3] != '-' else 0
				}
				data.append(item)

		return data

	@staticmethod
	def process_profit_loss_sheet(sheet):
		"""Обрабатывает лист с отчетом о прибылях и убытках"""
		data = []
		years = ['2013', '2014', '2015']

		# Парсим данные из таблицы
		for row in sheet.iter_rows(min_row=20, values_only=True):
			if not any(row):  # Пропускаем пустые строки
				continue

			if row[0] and isinstance(row[0], str) and row[2] and isinstance(row[2], int):
				# Это строка с данными
				item = {
					'parameter': row[0].strip(),
					'code': str(row[2]),
					'2013': float(row[5]) if row[5] and row[5] != '-' else 0,
					'2014': float(row[4]) if row[4] and row[4] != '-' else 0,
					'2015': float(row[3]) if row[3] and row[3] != '-' else 0
				}
				data.append(item)

		return data


class GraphDialog(QDialog):
	def __init__(self, parent=None):
		super().__init__(parent)
		self.setWindowTitle("График параметра")
		self.setMinimumSize(800, 600)

		# Создаем фигуру matplotlib
		self.figure = Figure(figsize=(8, 6), dpi=100)
		self.canvas = FigureCanvas(self.figure)

		# Настройка layout
		layout = QVBoxLayout()
		layout.addWidget(self.canvas)
		self.setLayout(layout)

	def plot_data(self, years, values, title, ylabel):
		"""Отрисовывает график на основе переданных данных"""
		self.figure.clear()
		ax = self.figure.add_subplot(111)

		# Отрисовываем график
		ax.plot(years, values, marker='o', linestyle='-', color='#4CAF50', linewidth=2, markersize=8)

		# Настройки графика
		ax.set_title(title, fontsize=14, pad=20)
		ax.set_xlabel('Год', fontsize=12)
		ax.set_ylabel(ylabel, fontsize=12)
		ax.grid(True, linestyle='--', alpha=0.7)
		ax.set_facecolor('#f0f0f0')

		# Форматирование осей
		ax.tick_params(axis='both', which='major', labelsize=10)

		# Автоматическое масштабирование
		ax.autoscale_view()

		# Обновляем холст
		self.canvas.draw()


class MainWindow(QMainWindow):
	def __init__(self):
		super().__init__()
		self.setWindowTitle("Анализ бухгалтерской отчетности")
		self.showMaximized()
		self.data = None  # Хранение всех данных
		self.current_year = 2015
		self.current_table = 'balance'  # 'balance' или 'profit_loss'
		self.updating_table = False  # Флаг для предотвращения рекурсии
		self.setup_ui()
		self.apply_styles()

		# Загружаем тестовые данные
		self.load_sample_data()

	def setup_ui(self):
		# Главный контейнер
		main_widget = QWidget()
		main_layout = QVBoxLayout(main_widget)
		main_layout.setContentsMargins(20, 20, 20, 20)
		main_layout.setSpacing(20)

		# Панель управления
		control_panel = QFrame()
		control_layout = QHBoxLayout(control_panel)
		control_layout.setContentsMargins(15, 15, 15, 15)
		control_layout.setSpacing(30)

		# Выбор года
		year_label = QLabel("Отчетный год:")
		self.year_combo = QComboBox()
		self.year_combo.addItems(["2013", "2014", "2015"])
		self.year_combo.currentTextChanged.connect(self.update_table)

		# Кнопки для выбора таблиц
		self.balance_btn = QPushButton("Бухгалтерский баланс")
		self.balance_btn.setCheckable(True)
		self.balance_btn.setChecked(True)
		self.balance_btn.clicked.connect(lambda: self.switch_table('balance'))

		self.profit_loss_btn = QPushButton("Отчет о прибылях и убытках")
		self.profit_loss_btn.setFixedWidth(200)
		self.profit_loss_btn.setCheckable(True)
		self.profit_loss_btn.clicked.connect(lambda: self.switch_table('profit_loss'))

		# Кнопка для показа графика
		self.show_graph_btn = QPushButton("Показать график")
		self.show_graph_btn.clicked.connect(self.show_graph)

		# Кнопка экспорта в Excel
		self.export_btn = QPushButton("Экспорт в Excel")
		self.export_btn.clicked.connect(self.export_to_excel)

		# Группа кнопок для взаимного исключения
		self.table_btn_group = QButtonGroup()
		self.table_btn_group.addButton(self.balance_btn)
		self.table_btn_group.addButton(self.profit_loss_btn)

		control_layout.addWidget(year_label)
		control_layout.addWidget(self.year_combo)
		control_layout.addStretch()
		control_layout.addWidget(self.balance_btn)
		control_layout.addWidget(self.profit_loss_btn)
		control_layout.addWidget(self.show_graph_btn)
		control_layout.addWidget(self.export_btn)
		# Настройка таблицы
		self.table = QTableWidget()
		self.table.setColumnCount(5)
		self.table.setHorizontalHeaderLabels([
			"Показатель",
			"Отчетный год",
			"Предыдущий год",
			"Темп роста, %",
			"Абсолютное отклонение"
		])
		self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
		self.table.verticalHeader().setVisible(False)


		# Разрешаем редактирование только столбцов с годами
		self.table.itemChanged.connect(self.handle_item_changed)

		# Настройка высоты строк
		self.table.verticalHeader().setDefaultSectionSize(30)
		self.table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

		# Добавление элементов
		main_layout.addWidget(control_panel)
		main_layout.addWidget(self.table)
		self.setCentralWidget(main_widget)

	def load_sample_data(self):
		"""Загружает тестовые данные"""
		self.data = {
			'balance': [
				{'section': 'I. ДОЛГОСРОЧНЫЕ АКТИВЫ'},
				{'parameter': 'Основные средства', 'code': '110', '2013': 275348, '2014': 220011, '2015': 293430},
				{'parameter': 'Нематериальные активы', 'code': '120', '2013': 1201, '2014': 560, '2015': 169},
				{'parameter': 'Доходные вложения в материальные активы', 'code': '130', '2013': 0, '2014': 0,
				 '2015': 0},
				{'parameter': 'Вложения в долгосрочные активы', 'code': '140', '2013': 50482, '2014': 51168,
				 '2015': 37388},
				{'parameter': 'Долгосрочные финансовые вложения', 'code': '150', '2013': 10, '2014': 0, '2015': 0},
				{'parameter': 'Долгосрочная дебиторская задолженность', 'code': '170', '2013': 0, '2014': 0,
				 '2015': 1500},
				{'parameter': 'Прочие долгосрочные активы', 'code': '180', '2013': 70395, '2014': 0, '2015': 101566},
				{'parameter': 'ИТОГО по разделу I', 'code': '190', '2013': 397436, '2014': 271739, '2015': 434053},
				{'section': 'II. КРАТКОСРОЧНЫЕ АКТИВЫ'},
				{'parameter': 'Запасы', 'code': '210', '2013': 73895, '2014': 67247, '2015': 54151},
				{'parameter': 'Расходы будущих периодов', 'code': '230', '2013': 841, '2014': 469, '2015': 202},
				{'parameter': 'Налог на добавленную стоимость', 'code': '240', '2013': 1939, '2014': 1965,
				 '2015': 2342},
				{'parameter': 'Краткосрочная дебиторская задолженность', 'code': '250', '2013': 56302, '2014': 25164,
				 '2015': 40230},
				{'parameter': 'Денежные средства и их эквиваленты', 'code': '270', '2013': 2523, '2014': 5044,
				 '2015': 4521},
				{'parameter': 'Прочие краткосрочные активы', 'code': '280', '2013': 15, '2014': 14, '2015': 14},
				{'parameter': 'ИТОГО по разделу II', 'code': '290', '2013': 135515, '2014': 99903, '2015': 101460},
				{'parameter': 'БАЛАНС', 'code': '300', '2013': 532951, '2014': 371642, '2015': 535513},
				{'section': 'III. СОБСТВЕННЫЙ КАПИТАЛ'},
				{'parameter': 'Уставный капитал', 'code': '410', '2013': 40255, '2014': 40255, '2015': 40255},
				{'parameter': 'Резервный капитал', 'code': '440', '2013': 1136, '2014': 1136, '2015': 1136},
				{'parameter': 'Добавочный капитал', 'code': '450', '2013': 231103, '2014': 70352, '2015': 111093},
				{'parameter': 'Нераспределенная прибыль (непокрытый убыток)', 'code': '460', '2013': -19182,
				 '2014': -57461, '2015': -101507},
				{'parameter': 'ИТОГО по разделу III', 'code': '490', '2013': 253312, '2014': 54282, '2015': 50977},
				{'section': 'IV. ДОЛГОСРОЧНЫЕ ОБЯЗАТЕЛЬСТВА'},
				{'parameter': 'Долгосрочные кредиты и займы', 'code': '510', '2013': 141710, '2014': 128580,
				 '2015': 128273},
				{'parameter': 'Доходы будущих периодов', 'code': '540', '2013': 177, '2014': 144, '2015': 132},
				{'parameter': 'ИТОГО по разделу IV', 'code': '590', '2013': 141887, '2014': 128724, '2015': 128405},
				{'section': 'V. КРАТКОСРОЧНЫЕ ОБЯЗАТЕЛЬСТВА'},
				{'parameter': 'Краткосрочные кредиты и займы', 'code': '610', '2013': 5397, '2014': 6992, '2015': 864},
				{'parameter': 'Краткосрочная часть долгосрочных обязательств', 'code': '620', '2013': 28940,
				 '2014': 99874, '2015': 239330},
				{'parameter': 'Краткосрочная кредиторская задолженность', 'code': '630', '2013': 103415, '2014': 81770,
				 '2015': 115937},
				{'parameter': 'ИТОГО по разделу V', 'code': '690', '2013': 137752, '2014': 188636, '2015': 356131},
				{'parameter': 'БАЛАНС', 'code': '700', '2013': 532951, '2014': 371642, '2015': 535513}
			],
			'profit_loss': [
				{'parameter': 'Выручка от реализации продукции, товаров, работ, услуг', 'code': '010',
				 '2013': 413323, '2014': 307922, '2015': 308637},
				{'parameter': 'Себестоимость реализованной продукции, товаров, работ, услуг', 'code': '020',
				 '2013': -383863, '2014': -294283, '2015': -256986},
				{'parameter': 'Валовая прибыль (010-020)', 'code': '030',
				 '2013': 29460, '2014': 13639, '2015': 51651},
				{'parameter': 'Управленческие расходы', 'code': '040',
				 '2013': -31904, '2014': -32102, '2015': -28369},
				{'parameter': 'Расходы на реализацию', 'code': '050',
				 '2013': -6879, '2014': -6489, '2015': -5616},
				{'parameter': 'Прибыль (убыток) от реализации продукции, товаров, работ, услуг (030-040-050)',
				 'code': '060', '2013': -9323, '2014': -24952, '2015': 17666},
				{'parameter': 'Прочие доходы по текущей деятельности', 'code': '070',
				 '2013': 244236, '2014': 153662, '2015': 146380},
				{'parameter': 'Прочие расходы по текущей деятельности', 'code': '080',
				 '2013': -248991, '2014': -154405, '2015': -146896},
				{'parameter': 'Прибыль (убыток) от текущей деятельности (060+070-080)', 'code': '090',
				 '2013': -14078, '2014': -25695, '2015': 17150},

				# Инвестиционная деятельность
				{'parameter': 'Доходы по инвестиционной деятельности', 'code': '100',
				 '2013': 1428, '2014': 2420, '2015': 1713},
				{'parameter': 'В том числе:', 'code': '', '2013': None, '2014': None, '2015': None, 'is_subitem': True},
				{
					'parameter': 'доходы от выбытия основных средств, нематериальных активов и других долгосрочных активов',
					'code': '101', '2013': 734, '2014': 676, '2015': 771, 'is_subitem': True},
				{'parameter': 'доходы от участия в уставном капитале других организаций',
				 'code': '102', '2013': 0, '2014': 0, '2015': 0, 'is_subitem': True},
				{'parameter': 'проценты к получению', 'code': '103',
				 '2013': 116, '2014': 192, '2015': 277, 'is_subitem': True},
				{'parameter': 'прочие доходы по инвестиционной деятельности', 'code': '104',
				 '2013': 578, '2014': 1552, '2015': 665, 'is_subitem': True},

				{'parameter': 'Расходы по инвестиционной деятельности', 'code': '110',
				 '2013': -819, '2014': -402, '2015': -213},
				{'parameter': 'В том числе:', 'code': '', '2013': None, '2014': None, '2015': None, 'is_subitem': True},
				{
					'parameter': 'расходы от выбытия основных средств, нематериальных активов и других долгосрочных активов',
					'code': '111', '2013': -800, '2014': -376, '2015': -189, 'is_subitem': True},
				{'parameter': 'прочие расходы по инвестиционной деятельности', 'code': '112',
				 '2013': -19, '2014': -26, '2015': -24, 'is_subitem': True},

				# Финансовая деятельность
				{'parameter': 'Доходы по финансовой деятельности', 'code': '120',
				 '2013': 9719, '2014': 4876, '2015': 10279},
				{'parameter': 'В том числе:', 'code': '', '2013': None, '2014': None, '2015': None, 'is_subitem': True},
				{'parameter': 'курсовые разницы от пересчета активов и обязательств', 'code': '121',
				 '2013': 2077, '2014': 0, '2015': 6239, 'is_subitem': True},
				{'parameter': 'прочие доходы по финансовой деятельности', 'code': '122',
				 '2013': 7642, '2014': 4876, '2015': 4040, 'is_subitem': True},

				{'parameter': 'Расходы по финансовой деятельности', 'code': '130',
				 '2013': -18917, '2014': -19894, '2015': -72916},
				{'parameter': 'В том числе:', 'code': '', '2013': None, '2014': None, '2015': None, 'is_subitem': True},
				{'parameter': 'проценты к уплате', 'code': '131',
				 '2013': -12215, '2014': -15326, '2015': -27814, 'is_subitem': True},
				{'parameter': 'курсовые разницы от пересчета активов и обязательств', 'code': '132',
				 '2013': -2077, '2014': 0, '2015': -42460, 'is_subitem': True},
				{'parameter': 'прочие расходы по финансовой деятельности', 'code': '133',
				 '2013': -4625, '2014': -4568, '2015': -2642, 'is_subitem': True},

				{'parameter': 'Прибыль (убыток) от инвестиционной, финансовой и иной деятельности (100-110+120-130)',
				 'code': '140', '2013': -8589, '2014': -13000, '2015': -61137},
				{'parameter': 'Прибыль (убыток) до налогообложения (090+140)', 'code': '150',
				 '2013': -22667, '2014': -38695, '2015': -43987},

				# Налоги и чистая прибыль
				{'parameter': 'Налог на прибыль', 'code': '160', '2013': 0, '2014': 0, '2015': 0},
				{'parameter': 'Изменение отложенных налоговых активов', 'code': '170', '2013': 0, '2014': 0, '2015': 0},
				{'parameter': 'Изменение отложенных налоговых обязательств', 'code': '180', '2013': 0, '2014': 0,
				 '2015': 0},
				{'parameter': 'Прочие налоги и сборы, исчисляемые из прибыли (дохода)', 'code': '190', '2013': 0,
				 '2014': 0, '2015': 0},
				{'parameter': 'Прочие платежи, исчисляемые из прибыли (дохода)', 'code': '200', '2013': 0, '2014': 0,
				 '2015': 0},

				{'parameter': 'Чистая прибыль (убыток) (150-160+170+180-190-200)', 'code': '210',
				 '2013': -22667, '2014': -38695, '2015': -43987},
				{'parameter': 'Количество прибыльных организаций', 'code': '211', '2013': 0, '2014': 0, '2015': 0},
				{'parameter': 'Сумма прибыли', 'code': '212', '2013': 0, '2014': 0, '2015': 0},
				{'parameter': 'Количество убыточных организаций', 'code': '213', '2013': 1, '2014': 1, '2015': 1},
				{'parameter': 'Сумма убытка', 'code': '214', '2013': -22667, '2014': -38695, '2015': -43987},

				{'parameter': 'Результат от переоценки долгосрочных активов, не включаемый в чистую прибыль (убыток)',
				 'code': '220', '2013': 43546, '2014': -42062, '2015': 40741},
				{'parameter': 'Результат от прочих операций, не включаемый в чистую прибыль (убыток)', 'code': '230',
				 '2013': 0, '2014': 0, '2015': 0},
				{'parameter': 'Совокупная прибыль (убыток) (210+220+230)', 'code': '240',
				 '2013': 20879, '2014': -80757, '2015': -3246},

				{'parameter': 'Базовая прибыль (убыток) на акцию, рублей', 'code': '250',
				 '2013': -1858, '2014': -3173, '2015': -3605},
				{'parameter': 'Разводненная прибыль (убыток) на акцию, рублей', 'code': '260',
				 '2013': 0, '2014': 0, '2015': 0}
			]
		}

		self.update_table()

	def handle_item_changed(self, item):
		"""Обрабатывает изменение значения в ячейке"""
		if self.updating_table or item.column() not in (1, 2):  # Только для столбцов с годами
			return

		selected_year = int(self.year_combo.currentText())
		prev_year = selected_year - 1 if selected_year > 2013 else None

		# Получаем данные строки
		row = item.row()
		parameter_item = self.table.item(row, 0)

		# Пропускаем строки с заголовками разделов
		if not parameter_item or parameter_item.text().startswith(('I.', 'II.', 'III.', 'IV.', 'V.')):
			return

		# Находим соответствующий параметр в данных
		current_data = self.data[self.current_table]
		data_item = None
		for item_data in current_data:
			if 'parameter' in item_data and item_data['parameter'] == parameter_item.text():
				data_item = item_data
				break

		if not data_item:
			return

		# Обновляем значение в данных
		try:
			new_value = float(item.text().replace(',', ''))
		except ValueError:
			QMessageBox.warning(self, "Ошибка", "Введите корректное числовое значение")
			self.update_table()
			return

		# Определяем, какой год изменяем
		if item.column() == 1:  # Отчетный год
			data_item[str(selected_year)] = new_value
		elif item.column() == 2 and prev_year:  # Предыдущий год
			data_item[str(prev_year)] = new_value

		# Если это строка в разделе, пересчитываем итоги раздела
		self.recalculate_section_totals()

		# Обновляем таблицу
		self.update_table()

	def recalculate_section_totals(self):
		"""Пересчитывает итоги по разделам"""
		selected_year = int(self.year_combo.currentText())
		prev_year = selected_year - 1 if selected_year > 2013 else None

		if self.current_table == 'balance':
			# Для бухгалтерского баланса
			sections = {
				'I.': ('190', ['110', '120', '130', '140', '150', '170', '180']),
				'II.': ('290', ['210', '230', '240', '250', '270', '280']),
				'III.': ('490', ['410', '440', '450', '460']),
				'IV.': ('590', ['510', '540']),
				'V.': ('690', ['610', '620', '630']),
				'BALANCE': ('700', ['190', '290', '490', '590', '690'])
			}

			for section, (total_code, item_codes) in sections.items():
				# Суммируем значения для текущего года
				total = 0
				for item in self.data['balance']:
					if 'code' in item and item['code'] in item_codes:
						total += item.get(str(selected_year), 0)

				# Обновляем итог
				for item in self.data['balance']:
					if 'code' in item and item['code'] == total_code:
						item[str(selected_year)] = total
						break

				# Суммируем значения для предыдущего года, если он есть
				if prev_year:
					prev_total = 0
					for item in self.data['balance']:
						if 'code' in item and item['code'] in item_codes:
							prev_total += item.get(str(prev_year), 0)

					# Обновляем итог
					for item in self.data['balance']:
						if 'code' in item and item['code'] == total_code:
							item[str(prev_year)] = prev_total
							break

	def export_to_excel(self):
		"""Экспортирует текущую таблицу с расчетами в Excel"""
		if not self.data:
			QMessageBox.warning(self, "Ошибка", "Нет данных для экспорта")
			return

		file_path, _ = QFileDialog.getSaveFileName(
			self, "Сохранить файл Excel", "", "Excel Files (*.xlsx)"
		)

		if file_path:
			try:
				workbook = openpyxl.Workbook()
				sheet = workbook.active

				# Устанавливаем название листа в зависимости от текущей таблицы
				if self.current_table == 'balance':
					sheet.title = "Бухгалтерский баланс"
				else:
					sheet.title = "Отчет о прибылях и убытках"

				# Заголовки столбцов
				headers = [
					"Показатель",
					"Отчетный год",
					"Предыдущий год",
					"Темп роста, %",
					"Абсолютное отклонение"
				]

				sheet.append(headers)

				# Настройка стилей для заголовков
				header_font = Font(bold=True, color="FFFFFF")
				header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
				header_alignment = Alignment(horizontal="center", vertical="center")

				for col in range(1, len(headers) + 1):
					cell = sheet.cell(row=1, column=col)
					cell.font = header_font
					cell.fill = header_fill
					cell.alignment = header_alignment

				# Получаем данные для экспорта
				current_data = self.data[self.current_table]
				selected_year = int(self.year_combo.currentText())
				prev_year = selected_year - 1 if selected_year > 2013 else None

				# Заполняем данные
				row_idx = 2  # Начинаем со второй строки
				current_section = None

				for item in current_data:
					if 'section' in item:
						# Это заголовок раздела
						current_section = item['section']
						sheet.append([current_section])

						# Объединяем ячейки для заголовка раздела
						sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)

						# Настройка стиля для заголовка раздела
						cell = sheet.cell(row=row_idx, column=1)
						cell.font = Font(bold=True)
						cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

						row_idx += 1
					else:
						# Это строка с данными
						current_val = item.get(str(selected_year), 0)
						prev_val = item.get(str(prev_year), 0) if prev_year else None

						# Рассчитываем темп роста
						if prev_val and prev_val != 0:
							growth = (current_val / prev_val) * 100
							growth_text = f"{growth:.2f}%"
						else:
							growth_text = "-"

						# Рассчитываем абсолютное отклонение
						if prev_val is not None:
							deviation = current_val - prev_val
							deviation_text = f"{deviation:,.2f}"
						else:
							deviation_text = "-"

						# Добавляем строку
						sheet.append([
							item['parameter'],
							current_val if current_val != 0 else "-",
							prev_val if prev_val and prev_val != 0 else "-",
							growth_text,
							deviation_text
						])

						# Настройка стилей для числовых значений
						for col in [2, 3, 5]:  # Столбцы с числами
							cell = sheet.cell(row=row_idx, column=col)
							cell.number_format = '#,##0.00'

						# Настройка цвета для отклонения
						if prev_val is not None:
							deviation_cell = sheet.cell(row=row_idx, column=5)
							if deviation < 0:
								deviation_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
																  fill_type="solid")
							else:
								deviation_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
																  fill_type="solid")

						row_idx += 1

				# Автонастройка ширины столбцов
				for column in sheet.columns:
					max_length = 0
					column_letter = get_column_letter(column[0].column)

					for cell in column:
						try:
							if len(str(cell.value)) > max_length:
								max_length = len(str(cell.value))
						except:
							pass

					adjusted_width = (max_length + 2) * 1.2
					sheet.column_dimensions[column_letter].width = adjusted_width

				# Сохраняем файл
				workbook.save(file_path)
				QMessageBox.information(self, "Успех", "Файл успешно экспортирован!")
			except Exception as e:
				QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать файл:\n{str(e)}")

	def show_graph(self):
		"""Показывает график выбранного параметра"""
		if not self.data:
			QMessageBox.warning(self, "Ошибка", "Нет данных для построения графика")
			return

		selected_year = int(self.year_combo.currentText())

		# Создаем диалоговое окно с графиком
		graph_dialog = GraphDialog(self)

		if self.current_table == 'balance':
			# График для бухгалтерского баланса
			# Выбираем параметр "ИТОГО по разделу I" (код 190)
			parameter_data = next((item for item in self.data['balance'] if item.get('code') == '190'), None)

			if parameter_data:
				years = ['2013', '2014', '2015']
				values = [parameter_data.get(year, 0) for year in years]

				graph_dialog.plot_data(
					years,
					values,
					"Динамика долгосрочных активов по годам",
					"Сумма, млн. руб."
				)
		else:
			# График для отчета о прибылях и убытках
			# Выбираем параметр "Чистая прибыль (убыток)" (код 210)
			parameter_data = next((item for item in self.data['profit_loss'] if item.get('code') == '210'), None)

			if parameter_data:
				years = ['2013', '2014', '2015']
				values = [parameter_data.get(year, 0) for year in years]

				graph_dialog.plot_data(
					years,
					values,
					"Динамика чистой прибыли (убытка) по годам",
					"Сумма, млн. руб."
				)

		graph_dialog.exec()

	def switch_table(self, table_name):
		self.current_table = table_name
		self.update_table()

	def update_table(self):
		if not self.data:
			return

		selected_year = int(self.year_combo.currentText())
		prev_year = selected_year - 1 if selected_year > 2013 else None

		# Очищаем таблицу перед заполнением
		self.table.clearContents()

		# Выбираем нужный набор данных в зависимости от текущей таблицы
		current_data = self.data[self.current_table]

		# Сначала подсчитаем общее количество строк с учетом разделов и подпунктов
		total_rows = 0
		for item in current_data:
			if 'section' in item:
				total_rows += 1  # Строка раздела
			else:
				total_rows += 1  # Основная строка
				# Если есть подпункты, добавляем их
				if item.get('is_subitem', False):
					total_rows += 1

		self.table.setRowCount(total_rows)

		current_section = None
		row_idx = 0

		# Устанавливаем флаг обновления таблицы
		self.updating_table = True

		for item in current_data:
			# Проверяем, нужно ли добавить разделитель раздела
			if 'section' in item and item['section'] != current_section:
				current_section = item['section']
				# Добавляем строку с названием раздела
				section_item = QTableWidgetItem(current_section)
				section_item.setFlags(section_item.flags() ^ Qt.ItemIsEditable)
				font = QFont()
				font.setBold(True)
				section_item.setFont(font)
				section_item.setBackground(QColor(70, 70, 70))
				self.table.setItem(row_idx, 0, section_item)
				# Объединяем ячейки для названия раздела
				self.table.setSpan(row_idx, 0, 1, 5)
				row_idx += 1

			if 'parameter' in item:
				# Заполняем строку данными
				# Показатель
				param_text = item['parameter']
				if item.get('is_subitem', False):
					param_text = "    " + param_text  # Добавляем отступ для подпунктов

				param_item = QTableWidgetItem(param_text)
				param_item.setFlags(param_item.flags() ^ Qt.ItemIsEditable)

				# Устанавливаем курсив для подпунктов
				if item.get('is_subitem', False):
					font = QFont()
					font.setItalic(True)
					param_item.setFont(font)

				self.table.setItem(row_idx, 0, param_item)

				# Отчетный год
				current_val = item.get(str(selected_year))
				if current_val is not None and current_val != 0:
					current_text = f"{current_val:,.2f}"
				else:
					current_text = "-"
				current_item = QTableWidgetItem(current_text)
				current_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
				self.table.setItem(row_idx, 1, current_item)

				# Предыдущий год
				if prev_year:
					prev_val = item.get(str(prev_year))
					if prev_val is not None and prev_val != 0:
						prev_text = f"{prev_val:,.2f}"
					else:
						prev_text = "-"
				else:
					prev_text = "-"
				prev_item = QTableWidgetItem(prev_text)
				prev_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
				self.table.setItem(row_idx, 2, prev_item)

				# Темп роста
				if prev_year and current_val is not None and prev_val is not None and prev_val != 0:
					growth = (current_val / prev_val) * 100
					growth_item = QTableWidgetItem(f"{growth:.2f}%")
				else:
					growth_item = QTableWidgetItem("-")
				growth_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
				self.table.setItem(row_idx, 3, growth_item)

				# Абсолютное отклонение
				if prev_year and current_val is not None and prev_val is not None:
					deviation = current_val - prev_val
					deviation_item = QTableWidgetItem(f"{deviation:,.2f}")
					# Раскраска
					if deviation < 0:
						deviation_item.setBackground(QColor(150, 50, 50))
					else:
						deviation_item.setBackground(QColor(50, 150, 50))
					deviation_item.setForeground(QColor(255, 255, 255))
				else:
					deviation_item = QTableWidgetItem("-")
				deviation_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
				self.table.setItem(row_idx, 4, deviation_item)

				row_idx += 1

		# Снимаем флаг обновления таблицы
		self.updating_table = False

	def apply_styles(self):
		self.setStyleSheet("""
            QMainWindow {
                background-color: #252525;
            }
            QFrame {
                background-color: #333333;
                border-radius: 8px;
            }
            QLabel {
                color: #ffffff;
                font-size: 14px;
            }
            QComboBox {
                background-color: #3c3c3c;
                color: #ffffff;
                border: 1px solid #555;
                border-radius: 5px;
                padding: 8px;
                min-width: 200px;
            }
            QComboBox QAbstractItemView {
                background-color: #3c3c3c;
                color: #ffffff;
                selection-background-color: #4CAF50;
            }
            QTableWidget {
                background-color: #333333;
                color: #ffffff;
                border: 1px solid #444;
                gridline-color: #444;
                font-size: 14px;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: #ffffff;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 6px;
            }
            QPushButton {
                background-color: #3c3c3c;
                color: #ffffff;
                border: 1px solid #555;
                border-radius: 5px;
                padding: 8px;
                min-width: 150px;
            }
            QPushButton:checked {
                background-color: #4CAF50;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #555;
            }
        """)


if __name__ == "__main__":
	from PySide6.QtWidgets import QApplication

	app = QApplication([])
	window = MainWindow()
	window.show()
	app.exec()
